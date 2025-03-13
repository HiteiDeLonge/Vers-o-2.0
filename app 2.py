# Este arquivo está sujeito aos termos e condições da Licença de Software Proprietário
# incluída no arquivo LICENSE.md que acompanha este software.

import sys
import tkinter as tk
from tkinter import Tk, Label, Button, ttk
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium import webdriver
from xml.etree import ElementTree as ET
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from xml.etree import ElementTree as ET
from lxml import html
import re
import pandas as pd
import pyautogui
import keyboard
import os
import zipfile
import rarfile
import shutil
from dateutil import parser
from tkinter import messagebox
from collections import defaultdict
import openpyxl
from PyPDF2 import PdfReader,PdfWriter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from itertools import islice
from selenium.webdriver.common.keys import Keys
import autoit
import win32com.client as win32
from tqdm import tqdm


#PREFERENCIAS DA PAGINA DO GASOLA

prefs = {"download.default_directory": r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas",
                 "safebrowsing.enabled": "true"}
    # Configurações do Chrome
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-features=ColorCorrectRendering")  # Desabilita o carregamento de cores
chrome_options.add_argument("--blink-settings=imagesEnabled=false")  # Desabilitar imagens
chrome_options.add_argument("--disable-gpu") #Desabilita GPU para melhor processamento
chrome_options.add_argument("--disable-web-security") # Desabilita a política de mesma origem
chrome_options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(options=chrome_options)

#INICIAR A PRIMEIRA TELA DE FECHAMENTOS:

def acessar_site():
    print("acessar_site()")
    driver.get("https://frota.gasola.net/gestao-financeira") 
    login_gasola_i()

def login_gasola_i():
    print("login_gasola_i()")   
    #escrevendo usuário
    input_user = driver.find_elements(By.XPATH, '//*[@id="root"]/div[3]/div/div/div[1]/input')[0]
    input_user.send_keys("AQUI VAI O USUÁRIO")
    #escrevendo senha
    input_pass = driver.find_elements(By.XPATH, '//*[@id="root"]/div[3]/div/div/div[2]/div/input')[0]
    input_pass.send_keys("E AQUI VAI A SENHA :) ")

    # Clicando no botão de login
    login_button = driver.find_elements(By.XPATH, '//*[@id="root"]/div[3]/div/div/button[2]')[0]
    login_button.click()
    sleep(2)
    filtrar_nao_pagas_i()

def filtrar_nao_pagas_i():
    print("filtrar_nao_pagas_i()")
    sleep(3)
    botao_filtro = driver.find_elements(By.XPATH,'//*[@id="root"]/main/div[2]/div[1]/div/div[5]/div/div/div[1]')[0]
    botao_filtro.click()
    filtrar_nao = driver.find_elements(By.XPATH,'//*[@id="react-select-5-option-2"]')[0]
    filtrar_nao.click()
    botao_filtrar = driver.find_elements(By.XPATH,'//*[@id="root"]/main/div[2]/div[1]/div/button[2]')[0]
    botao_filtrar.click()
    sleep(1)
    captura_informacoes()

def verifica_licenca_presente():
  if os.path.isfile("LICENSE.md"):
    return True
  else:
    messagebox.showerror(
        "Licença não encontrada",
        "O arquivo de LICENSE.txt não foi encontrado.\nÉ proibido o uso, cópia, redistribuição e modificação deste software, total ou parcialmente, por qualquer pessoa ou entidade que não seja expressamente autorizada por escrito por Bernardo Augusto Borges Vilbert."
    )
    sys.exit()

def captura_informacoes():
    print("captura_informacoes()")
    # Obter o HTML da página
    html_content = driver.page_source
    # Parsear o HTML
    tree = html.fromstring(html_content)

    valores_codigo = []
    valores_nome = []
    valorp_posto = []
    valor_periodo_inicio = []
    valor_periodo_fim = []
    valor_quantidade_arquivos = []

    for i in range(1,11):
        codigo_elemento = tree.xpath('//*[@id="root"]/main/div[2]/div[2]/div[1]/div/div/table[{}]/thead/tr/th[1]'.format(i))
        if codigo_elemento:
            codigo_valor = codigo_elemento[0].text.strip() if codigo_elemento[0].text else None
            # Aqui é onde a modificação ocorre, apenas a parte numérica é mantida
            codigo_Número = codigo_valor.split(": ")[1].split(")")[0] if codigo_valor else None
            valores_codigo.append(codigo_Número)
        else:
            valores_codigo.append("")

        nome_elemento = driver.find_elements(By.XPATH,'//*[@id="root"]/main/div[2]/div[2]/div[1]/div/div/table[{}]/tbody/tr/td[1]/div/div[1]/div/p'.format(i))
        if nome_elemento:
            nome_valor = nome_elemento[0].text.strip() if nome_elemento[0].text else None
            valores_nome.append(nome_valor)
        else:
            valores_nome.append("")
            
        valor_elemento = driver.find_elements(By.XPATH,'//*[@id="root"]/main/div[2]/div[2]/div[1]/div/div/table[{}]/tbody/tr/td[7]/div/div/p[2]'.format(i))
        if valor_elemento:
            valor_valor = valor_elemento[0].text.strip() if valor_elemento[0].text else None
            valorp_posto.append(valor_valor)
        else:
            valorp_posto.append("")

        periodo_elemento = driver.find_elements(By.XPATH,'//*[@id="root"]/main/div[2]/div[2]/div[1]/div/div/table[{}]/tbody/tr/td[1]/div/div[5]'.format(i))
        if periodo_elemento:
            periodo_valor = periodo_elemento[0].text.strip() if periodo_elemento[0].text else None
            # Usando expressões regulares para extrair as datas de início e fim
            match = re.search(r'(\d{2}/\d{2}/\d{4}) \d{2}:\d{2} a (\d{2}/\d{2}/\d{4}) \d{2}:\d{2}', periodo_valor)
            if match:
                inicio, fim = match.groups()
                valor_periodo_inicio.append(inicio)
                valor_periodo_fim.append(fim)
            else:
                valor_periodo_inicio.append("00/00/0000")
                valor_periodo_fim.append("00/00/0000")
        else:
            valor_periodo_inicio.append("")
            valor_periodo_fim.append("")

        quantidade_arquivos_elemento = driver.find_elements(By.XPATH,'//*[@id="root"]/main/div[2]/div[2]/div[1]/div/div/table[{}]/tbody/tr/td[1]/div/div[7]/button'.format(i))
        if quantidade_arquivos_elemento:
            quantidade_arquivos_valor = quantidade_arquivos_elemento[0].text.strip() if quantidade_arquivos_elemento[0].text else None
            valor_quantidade_arquivos.append(quantidade_arquivos_valor)
        else:
            valor_quantidade_arquivos.append("")

    lista_zipada = list(zip(valores_codigo, valores_nome, valorp_posto, valor_periodo_inicio, valor_periodo_fim, valor_quantidade_arquivos))

    # Criar um novo DataFrame vazio
    df_vazio = pd.DataFrame()

    # Salvando o DataFrame vazio em um arquivo Excel
    nome_arquivo = r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\fechamentos.xlsx"
    df_vazio.to_excel(nome_arquivo, index=False)

    # Salvando os dados em um DataFrame
    df = pd.DataFrame(lista_zipada, columns=['Código', 'Nome', 'Valor', 'Período Início', 'Período Fim', 'Quantidade de Arquivos'])
    # Salvando o DataFrame em um arquivo Excel
    nome_arquivo = r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\fechamentos.xlsx"
    df.to_excel(nome_arquivo, index=False)

    return lista_zipada  # Opcional: retornar a lista zipada

def atualizar_tabela():
    print("atualizar_tabela()")
    dados = captura_informacoes()
    # Limpa a tabela existente
    for item in tree.get_children():
        tree.delete(item)

    # Adiciona os novos dados na tabela
    for dado in dados:
        tree.insert('', 'end', values=dado)

def centralizar_janela(janela):
    largura_tela = janela.winfo_screenwidth()
    altura_tela = janela.winfo_screenheight()

    largura_janela = 800  # Defina a largura da sua janela aqui
    altura_janela = 500   # Defina a altura da sua janela aqui

    x = (largura_tela - largura_janela) // 2
    y = (altura_tela - altura_janela) // 2

    janela.geometry("{}x{}+{}+{}".format(largura_janela, altura_janela, x, y))

def popup_menu(event):
    # Verifica se algum item da Treeview foi clicado
    item_id = tree.identify_row(event.y)
    if item_id:
        # Obtém as informações do item selecionado
        item_info = tree.item(item_id)
        # Obtém o valor da primeira coluna ("valores_codigo") do item selecionado
        codigo_find = item_info['values'][0]
        # Cria o menu pop-up
        menu = tk.Menu(janela, tearoff=0)
        fonte = ("Segoe UI", 10)  # Especificando apenas o nome e o tamanho da fonte
        # Adiciona a opção "Rateio" ao menu pop-up
        menu.add_command(label="Preparar Rateio     ", background="#99A1A4",foreground="#ffffff",font=fonte,command=lambda: buscar_fechamento_rateio(codigo_find))
        menu.add_command(label="Preparar Nota Placa ", background="#99A1A4",foreground="#ffffff",font=fonte,command=lambda: buscar_fechamento_notaplaca(codigo_find))
        menu.add_command(label="Iniciar NotaPlaca   ",background="#99A1A4",foreground="#ffffff",font=fonte,command=nota_placa)
        menu.add_command(label="Iniciar Rateio      ",background="#99A1A4",foreground="#ffffff",font=fonte,command=rateio)
        menu.add_command(label="Atualizar           ",background="#99A1A4",foreground="#ffffff",font=fonte,command=botao_atualizar)
        # Exibe o menu pop-up nas coordenadas do cursor do mouse
        menu.post(event.x_root, event.y_root)

def botao_atualizar():
    botao_filtrar = driver.find_elements(By.XPATH,'//*[@id="root"]/main/div[2]/div[1]/div/button[2]')[0]
    botao_filtrar.click()
    sleep(3)
    captura_informacoes()
    atualizar_tabela()
    messagebox.showinfo("Atualizando...","Fechamentos atualizados com sucesso")

#FUNÇÕES DE NOTA PLACA

def buscar_fechamento_notaplaca(codigo_find):
    print("buscar_fechamento_notaplaca(codigo_find)(Nota Placa)")
    global cont_nota
    limpar_pasta_notas()
    # Obter o HTML da página
    html_content = driver.page_source
    # Parsear o HTML
    tree = html.fromstring(html_content)
    fechamento_encontrado = False
    for i in range(1, 11):
        codigo_elemento = tree.xpath('//*[@id="root"]/main/div[2]/div[2]/div[1]/div/div/table[{}]/thead/tr/th[1]'.format(i))
        if codigo_elemento:
            codigo_valor = codigo_elemento[0].text.strip() if codigo_elemento[0].text else None
            # Aqui é onde a modificação ocorre, apenas a parte numérica é mantida
            codigo_Número = codigo_valor.split(": ")[1].split(")")[0] if codigo_valor else None
            if str(codigo_find) == codigo_Número:
                xpath_botao = '//*[@id="root"]/main/div[2]/div[2]/div[1]/div/div/table[{}]/tbody/tr/td[1]/div/div[7]/button'.format(i)
                botao = driver.find_element(By.XPATH, xpath_botao)
                # Clicar no botão para baixar
                botao.click()
                sleep(3)
                fechamento_encontrado = True
                extrair_e_excluir_arquivo()
                pasta_xmls = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas\xmls NF-e"
                for nome_arquivo in os.listdir(pasta_xmls):
                    caminho_xml = os.path.join(pasta_xmls, nome_arquivo)
                    resultados = processar_xml(caminho_xml, planilha_de_placas)

                    resultados_totais = []  # Lista para armazenar todos os resultados processados

                    for nome_arquivo in os.listdir(pasta_xmls):
                        caminho_xml = os.path.join(pasta_xmls, nome_arquivo)
                        resultados = processar_xml(caminho_xml, planilha_de_placas)
                        resultados_totais.extend(resultados)  # Adiciona os resultados ao total

                # Criar um DataFrame com todos os resultados
                df_resultados = pd.DataFrame(resultados_totais)

                # Adicionar os cabeçalhos ao DataFrame
                df_resultados.columns = ["Número", "Série", "Emitente CNPJ", "Emissão", "Descrição", "Qnt", "Reais/L", "Total", "Placa", "Emitente", "Filial"]
                
                # Salvar o DataFrame em um arquivo XLSX
                caminho_arquivo_xlsx = r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\notas.xlsx"
                df_resultados.to_excel(caminho_arquivo_xlsx, index=False)

                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(["NF", "Prod", "ICMS"])

                for icms in resultados_totais:
                    if "S10" in icms[4] or "S500" in icms[4]:
                        # Fazer a multiplicação do elemento 5 por 0.9456
                        nf_icms = float(icms[0])
                        prod_icms = (icms[4])
                        qnt_diesel = float(icms[5].replace(',', '.'))
                        resultado_icms = qnt_diesel * 0.9456
                        resultado_icms = round(resultado_icms, 2)  # Arredonda para 2 casas decimais
                        sheet.append([nf_icms, prod_icms, resultado_icms])
                caminho_icms = os.path.join(os.getcwd(), r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\icms.xlsx")
                workbook.save(caminho_icms)


                ultimo_item = resultados_totais[-1]
                mensagem = f"{ultimo_item[10]}"
                messagebox.showinfo("Lançar na Filial:",mensagem)
                
                break
    if not fechamento_encontrado:
        messagebox.showinfo("Não encontrado","Fechamento não entrado")
    verificar_pasta_notas_impressao()

def verificar_pasta_notas_impressao():
    caminho = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas\xmls NF-e"
    # Verificar o número de arquivos no diretório
    num_arquivos = len(os.listdir(caminho))
    if num_arquivos >= 20:
        messagebox.showerror("Convertendo XML em PDF", "Ops, parece que você excedeu o limite de arquivos.\nInfelizmente você terá que converter manualmente.")
        deletar_notas_lancadas()
        excluir_arquivo_notaplaca()
    else:
        converter_xml_np_impressao()
        extrair_e_excluir_arquivo()
        renomear_e_enviar_arquivo_impressao()
        deletar_notas_lancadas()
        excluir_arquivo_notaplaca()

def verificar_pasta_notas_fiscal():
    caminho = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas\xmls NF-e"
    # Verificar o número de arquivos no diretório
    num_arquivos = len(os.listdir(caminho))
    if num_arquivos >= 20:
        messagebox.showerror("Convertendo XML em PDF", "Ops, parece que você excedeu o limite de arquivos.\nInfelizmente você terá que converter manualmente.")
        verificar_notas_planilha()
    else:
        converter_xml_np()
        verificar_notas_planilha()
        
def converter_xml_np():
    print("converter_xml_np(Nota Placa)")
    # Defina o caminho para o seu diretório de download
    caminho_download = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas"
    
    # Configuração das opções do Chrome para definir o diretório de download
    options2 = Options()
    prefs = {"download.default_directory": caminho_download}
    options2.add_argument("--disable-features=ColorCorrectRendering")  # Desabilita o carregamento de cores
    options2.add_argument("--blink-settings=imagesEnabled=false")  # Desabilitar imagens
    options2.add_argument("--disable-gpu") #Desabilita GPU para melhor processamento
    options2.add_argument("--disable-web-security") # Desabilita a política de mesma origem
    options2.add_experimental_option("prefs", prefs)
    
    # Inicializa o driver do Chrome com as opções configuradas
    driver2 = webdriver.Chrome(options=options2)
    driver2.get("https://www.fsist.com.br/converter-xml-nfe-para-danfe")

    click_select = driver2.find_element(By.XPATH,'//*[@id="arquivolab"]') 
    click_select.click()
    
    autoit.win_wait_active("Abrir", 5)
    caminho = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas\xmls NF-e"
    
    # Define diretamente o texto do campo de edição com o caminho do diretório
    autoit.control_set_text("Abrir", "Edit1", caminho)
    autoit.control_send("Abrir", "Edit1", "{ENTER}")

  # Listar todos os arquivos dentro da pasta caminho e formatar apenas os arquivos .xml
    arquivos = os.listdir(caminho)
    nomes_formatados = ""
    for arquivo in arquivos:
        nome, extensao = os.path.splitext(arquivo)
        if extensao == ".xml":
            nomes_formatados += f'"{nome}" '

    sleep(2)
    autoit.control_send("Abrir", "Edit1", nomes_formatados.strip())
    sleep(1)
    autoit.control_send("Abrir", "Edit1", "{ENTER}")
    sleep(1)
    botao_gerar = driver2.find_element(By.XPATH,'//*[@id="divPlaceHolder"]/div[1]/table[1]/tbody/tr[1]/td[2]/label')
    botao_gerar.click()
    botao_confirmar = driver2.find_element(By.XPATH,'//*[@id="msgsim"]')
    botao_confirmar.click()
    sleep(1)
    botao_download = driver2.find_element(By.XPATH,'//*[@id="butlinktexto"]')
    botao_download.click()
    sleep(2)
    driver2.quit()
    sleep(2)
    extrair_e_excluir_arquivo()
    renomear_e_enviar_arquivo()
    encontrar_nota_em_pdf(diretorio_projeto)

def extrair_e_excluir_arquivo():
    print("extrair_e_excluir_arquivo(Nota Placa)")
    # Diretório onde os arquivos foram baixados
    diretorio_notas = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas"
    # Obter lista de arquivos baixados
    lista_arquivos = os.listdir(diretorio_notas)
    for arquivo in lista_arquivos:
        caminho_arquivo = os.path.join(diretorio_notas, arquivo)
        try:
            if arquivo.endswith('.zip'):
                # Descompacta o arquivo ZIP
                with zipfile.ZipFile(caminho_arquivo, 'r') as zip_ref:
                    zip_ref.extractall(diretorio_notas)
                # Exclui o arquivo compactado
                os.remove(caminho_arquivo)
            elif arquivo.endswith('.rar'):
                # Descompacta o arquivo RAR
                with rarfile.RarFile(caminho_arquivo, 'r') as rar_ref:
                    rar_ref.extractall(diretorio_notas)
                # Exclui o arquivo compactado
                os.remove(caminho_arquivo)
        except PermissionError as e:
            print(f"Erro de permissão ao excluir o arquivo {caminho_arquivo}: {e}")
        except Exception as e:
            print(f"Erro ao processar o arquivo {caminho_arquivo}: {e}")

cont_nota = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas\xmls NF-e"

def contar_arquivos_notaplaca(cont_nota):
    print("contar_arquivos_notaplaca(Nota Placa)")
    arquivos_da_pasta = os.listdir(cont_nota)
    num_arquivos = len(arquivos_da_pasta)
    if num_arquivos >=1:
        print(num_arquivos)
        verificar_pasta_notas_fiscal()
    else:
        messagebox.showinfo("Convertendo arquivos para o fiscal","Nenhuma nota para entregar para o fiscal")

def excluir_arquivo_notaplaca():
    print("excluir_arquivo_notaplaca(Nota Placa)")
    caminho_arquivo = r'C:\Users\Log20-2\Desktop\Versão 2.0\modules\excluir xml notaplaca.py'
    try:
        os.startfile(caminho_arquivo)
        sleep(3)
        contar_arquivos_notaplaca(cont_nota)
    except FileNotFoundError:
        print("O arquivo não foi encontrado.")

def deletar_notas_lancadas():
    print("deletar_notas_lancadas(Nota Placa)")
    base_conferencia = r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\notas.xlsx"
    
    try:
        conferencia_df = pd.read_excel(r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\conferencia.xlsx")
    except FileNotFoundError:
        print("O arquivo 'conferencia.xlsx' não pôde ser encontrado.")
        return
    
    notas_df = pd.read_excel(base_conferencia)
    
    # Converter células da coluna 'Número' para números
    notas_df['Número'] = pd.to_numeric(notas_df['Número'], errors='coerce')
    
    # Verificar notas e datas
    notas_e_datas_conferencia = conferencia_df[['NOTA', 'DATA']]
    notas_e_datas_notas = notas_df[['Número', 'Emissão']]
    
    notas_conferencia = notas_e_datas_conferencia.set_index('NOTA').to_dict()['DATA']
    notas_notas = notas_e_datas_notas.set_index('Número').to_dict()['Emissão']
    
    notas_a_remover = []
    for nota, emissao in notas_notas.items():
        if nota in notas_conferencia and notas_conferencia[nota] == emissao:
            notas_a_remover.append(nota)
    
    if notas_a_remover:
        notas_df = notas_df[~notas_df['Número'].isin(notas_a_remover)]
        notas_df.to_excel(base_conferencia, index=False)
        print("Notas removidas com sucesso.")
    else:
        print("Nenhuma nota encontrada para remover.")

def verificar_notas_planilha():
    caminho_pendencias = r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\notas.xlsx"
    verificar_notas_nao_lancadas = pd.read_excel(caminho_pendencias)
    linhas_com_numeros = verificar_notas_nao_lancadas[verificar_notas_nao_lancadas['Número'] != 'Título do Cabeçalho']
    if len(linhas_com_numeros) == 0:
        messagebox.showerror("ATENÇÃO", "Todas as notas já estão lançadas")
    else:
        verificar_placas_np()

def verificar_placas_np():
    caminho_planilha = r'C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\notas.xlsx'
    
    try:
        # Carregar a planilha
        df = pd.read_excel(caminho_planilha)
        
        # Verificar se alguma linha da coluna "Placa" possui os valores especificados
        placa_nao_encontrada = df[df['Placa'].isin(["Placa não encontrada na planilha", "Placa não encontrada no campo de observações"])]
        
        if not placa_nao_encontrada.empty:
            quantidade = len(placa_nao_encontrada)
            print(f"Foram encontradas {quantidade} placas não encontradas:")
            abrir_planilha_notas()
        else:
            messagebox.showinfo("Preparando Nota Placa","Fechamento pronto para ser lançado")
    
    except FileNotFoundError:
        print(f"O arquivo '{caminho_planilha}' não foi encontrado.")
    except Exception as e:
        print("Ocorreu um erro ao processar a planilha:", e)

def abrir_planilha_notas():
    caminho_arquivo = r'C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\notas.xlsx'
    try:
        os.startfile(caminho_arquivo)
        print("Arquivo aberto com sucesso!")
    except FileNotFoundError:
        print("O arquivo não foi encontrado.")

def renomear_e_enviar_arquivo():
    print("renomear_e_enviar_arquivo(Nota Placa)")
    # Diretório de origem e destino
    diretorio_origem = r'C:\Users\Log20-2\Desktop\Versão 2.0\Notas'
    diretorio_destino = r'C:\Users\Log20-2\Desktop\Versão 2.0'
    
    # Verifica se o diretório de origem existe
    if not os.path.exists(diretorio_origem):
        print(f"O diretório de origem '{diretorio_origem}' não existe.")
        return
    
    # Lista todos os arquivos no diretório de origem
    arquivos = os.listdir(diretorio_origem)
    
    # Verifica se o arquivo _JUNTO.pdf existe no diretório de origem
    if '_JUNTO.pdf' not in arquivos:
        print("O arquivo '_JUNTO.pdf' não foi encontrado no diretório de origem.")
        return
    
    # Obtém o caminho completo do arquivo _JUNTO.pdf
    arquivo_origem = os.path.join(diretorio_origem, '_JUNTO.pdf')
    
    # Move o arquivo para o diretório de destino com o nome 'notas.pdf'
    novo_nome = os.path.join(diretorio_destino, 'notas.pdf')
    shutil.move(arquivo_origem, novo_nome)
    
def encontrar_nota_em_pdf(diretorio_projeto):
    print("encontrar_nota_em_pdf(Nota Placa)")
    arquivo_excel = os.path.join(diretorio_projeto, r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\icms.xlsx")
    arquivo_pdf = os.path.join(diretorio_projeto, "notas.pdf")

    wb = openpyxl.load_workbook(arquivo_excel)
    planilha = wb.active

    writer = PdfWriter()

    notas_processadas = set()

    with open(arquivo_pdf, 'rb') as pdf_file:
        pdf_reader = PdfReader(pdf_file)

        # Dicionário para armazenar os produtos associados a cada nota fiscal
        produtos_por_nf = {}

        for linha in islice(planilha.iter_rows(values_only=True), 1, None):
            numero_nf, prod, icms = linha[:3]
            numero_nf_formatado = formatar_numero_nf(numero_nf)

            # Adiciona o produto à lista correspondente à nota fiscal
            produtos_por_nf.setdefault(numero_nf_formatado, []).append((prod, icms))

        for numero_nf_formatado, produtos in produtos_por_nf.items():
            # Verifica se a nota já foi processada
            if numero_nf_formatado in notas_processadas:
                continue

            encontrado = False

            for pagina_num, pagina in enumerate(pdf_reader.pages, 1):
                texto = pagina.extract_text()
                if numero_nf_formatado in texto:
                    # Cria uma página com a marca d'água para a nota fiscal
                    adicionar_marca_dagua(produtos)
                    pagina_com_marca = PdfReader("temp.pdf").pages[0]
                    # Adiciona marca d'água para cada produto associado à nota fiscal
                    pagina_com_marca.merge_page(pagina)
                    writer.add_page(pagina_com_marca)
                    encontrado = True
                    break

            notas_processadas.add(numero_nf_formatado)

        # Adiciona as páginas originais das notas não processadas
        for pagina in pdf_reader.pages:
            texto = pagina.extract_text()
            if all(nf not in texto for nf in notas_processadas):
                writer.add_page(pagina)

    with open(os.path.join(diretorio_projeto, "notas para o fiscal.pdf"), 'wb') as output_pdf:
        writer.write(output_pdf)

    # Verifica se o arquivo temp.pdf existe
    if os.path.exists("temp.pdf"):
        # Se existir, remove o arquivo
        os.remove("temp.pdf")
    else:
        # Se não existir, exibe uma mensagem
        messagebox.showinfo("Verifique o PDF de notas","Verifique o PDF de notas")

def adicionar_marca_dagua(valores):
    print("adicionar_marca_dagua(NOTA PLACA)")
    overlay = canvas.Canvas("temp.pdf", pagesize=letter)
    overlay.setFont("Helvetica", 16)
    y = 100
    for prod, icms in valores:
        overlay.drawString(100, y, f"Produto: {prod}, ICMS: {icms}")
        y += 20
    overlay.save()  

def converter_xml():
    print("converter_xml(Nota Placa)")
    # Defina o caminho para o seu diretório de download
    caminho_download = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas"
    
    # Configuração das opções do Chrome para definir o diretório de download
    options2 = Options()
    prefs = {"download.default_directory": caminho_download}
    options2.add_argument("--disable-features=ColorCorrectRendering")  # Desabilita o carregamento de cores
    options2.add_argument("--blink-settings=imagesEnabled=false")  # Desabilitar imagens
    options2.add_argument("--disable-gpu") #Desabilita GPU para melhor processamento
    options2.add_argument("--disable-web-security") # Desabilita a política de mesma origem
    options2.add_experimental_option("prefs", prefs)
    
    # Inicializa o driver do Chrome com as opções configuradas
    driver2 = webdriver.Chrome(options=options2)
    driver2.get("https://www.fsist.com.br/converter-xml-nfe-para-danfe")

    click_select = driver2.find_element(By.XPATH,'//*[@id="arquivolab"]') 
    click_select.click()
    
    autoit.win_wait_active("Abrir", 5)
    caminho = "C:\\Users\\Log20-2\\Desktop\\Versão 2.0\\Notas"
    
    # Define diretamente o texto do campo de edição com o caminho do diretório
    autoit.control_set_text("Abrir", "Edit1", caminho)
    autoit.control_send("Abrir", "Edit1", "{ENTER}")

  # Listar todos os arquivos dentro da pasta caminho e formatar apenas os arquivos .xml
    arquivos = os.listdir(caminho)
    nomes_formatados = ""
    for arquivo in arquivos:
        nome, extensao = os.path.splitext(arquivo)
        if extensao == ".xml":
            nomes_formatados += f'"{nome}" '

    sleep(2)
    autoit.control_send("Abrir", "Edit1", nomes_formatados.strip())
    sleep(1)
    autoit.control_send("Abrir", "Edit1", "{ENTER}")
    sleep(1)
    botao_gerar = driver2.find_element(By.XPATH,'//*[@id="divPlaceHolder"]/div[1]/table[1]/tbody/tr[1]/td[2]/label')
    botao_gerar.click()
    botao_confirmar = driver2.find_element(By.XPATH,'//*[@id="msgsim"]')
    botao_confirmar.click()
    sleep(1)
    botao_download = driver2.find_element(By.XPATH,'//*[@id="butlinktexto"]')
    botao_download.click()
    sleep(2)
    driver2.quit()

def converter_xml_np_impressao():
    print("converter_xml_np_impressao(Nota Placa)")
    # Defina o caminho para o seu diretório de download
    caminho_download = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas"
    
    # Configuração das opções do Chrome para definir o diretório de download
    options2 = Options()
    prefs = {"download.default_directory": caminho_download}
    options2.add_argument("--disable-features=ColorCorrectRendering")  # Desabilita o carregamento de cores
    options2.add_argument("--blink-settings=imagesEnabled=false")  # Desabilitar imagens
    options2.add_argument("--disable-gpu") #Desabilita GPU para melhor processamento
    options2.add_argument("--disable-web-security") # Desabilita a política de mesma origem
    options2.add_experimental_option("prefs", prefs)
    
    # Inicializa o driver do Chrome com as opções configuradas
    driver2 = webdriver.Chrome(options=options2)
    driver2.get("https://www.fsist.com.br/converter-xml-nfe-para-danfe")

    click_select = driver2.find_element(By.XPATH,'//*[@id="arquivolab"]') 
    click_select.click()
    
    autoit.win_wait_active("Abrir", 5)
    caminho = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas\xmls NF-e"
    
    # Define diretamente o texto do campo de edição com o caminho do diretório
    autoit.control_set_text("Abrir", "Edit1", caminho)
    autoit.control_send("Abrir", "Edit1", "{ENTER}")

  # Listar todos os arquivos dentro da pasta caminho e formatar apenas os arquivos .xml
    arquivos = os.listdir(caminho)
    nomes_formatados = ""
    for arquivo in arquivos:
        nome, extensao = os.path.splitext(arquivo)
        if extensao == ".xml":
            nomes_formatados += f'"{nome}" '
    print(nomes_formatados)
    sleep(2)
    autoit.control_send("Abrir", "Edit1", nomes_formatados.strip())
    sleep(1)
    autoit.control_send("Abrir", "Edit1", "{ENTER}")
    sleep(1)
    botao_gerar = driver2.find_element(By.XPATH,'//*[@id="divPlaceHolder"]/div[1]/table[1]/tbody/tr[1]/td[2]/label')
    botao_gerar.click()
    botao_confirmar = driver2.find_element(By.XPATH,'//*[@id="msgsim"]')
    botao_confirmar.click()
    sleep(1)
    botao_download = driver2.find_element(By.XPATH,'//*[@id="butlinktexto"]')
    botao_download.click()
    sleep(2)
    driver2.quit()
    
def renomear_e_enviar_arquivo_impressao():
    print("renomear_e_enviar_arquivo_impressao(Nota Placa)")
    # Diretório de origem e destino
    diretorio_origem = r'C:\Users\Log20-2\Desktop\Versão 2.0\Notas'
    diretorio_destino = r'C:\Users\Log20-2\Desktop\Versão 2.0'
    
    # Verifica se o diretório de origem existe
    if not os.path.exists(diretorio_origem):
        print(f"O diretório de origem '{diretorio_origem}' não existe.")
        return
    
    # Lista todos os arquivos no diretório de origem
    arquivos = os.listdir(diretorio_origem)
    
    # Verifica se o arquivo _JUNTO.pdf existe no diretório de origem
    if '_JUNTO.pdf' not in arquivos:
        print("O arquivo '_JUNTO.pdf' não foi encontrado no diretório de origem.")
        return
    
    # Obtém o caminho completo do arquivo _JUNTO.pdf
    arquivo_origem = os.path.join(diretorio_origem, '_JUNTO.pdf')
    
    # Move o arquivo para o diretório de destino com o nome 'notas.pdf'
    novo_nome = os.path.join(diretorio_destino, 'Notas para impressão.pdf')
    shutil.move(arquivo_origem, novo_nome)

# Caminho do arquivo CSV e caminho para salvar o arquivo Excel
arquivo_csv = r'C:\Users\Log20-2\Desktop\Versão 2.0\Notas\abastecimentos.csv'
caminho_salvamento = r'C:\Users\Log20-2\Desktop\Versão 2.0\Rateio\abastecimentos.xlsx'

#FUNÇÕES LANÇAMENTO NOTAPLACA:
def nota_placa():
    caminho_excel = r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\notas.xlsx"
    df = pd.read_excel(caminho_excel)

    # Agrupar por informação da nota fiscal
    df_agrupado = df.groupby(['Número', 'Série', 'Emitente CNPJ', 'Emissão'])

    for chave, grupo_df in df_agrupado:
        # Imprimir cabeçalho para cada nota fiscal
        imprimir_cabecalho_nota(grupo_df.iloc[0])

        # Iterar pelos itens da nota fiscal atual
        for _, row in grupo_df.iterrows():
            imprimir_item_nota(row)

            if keyboard.is_pressed('q'):
                break
        pyautogui.press('esc')    

def imprimir_cabecalho_nota(row):
    nfe = str(row['Número'])
    serie = str(row['Série'])
    modelo = "55"
    cnpj = str(row['Emitente CNPJ'])
    data = str(row['Emissão']).replace('/', '')
    operacao = "5002"

    pyautogui.click(1492, 89, duration=0.2)
    pyautogui.click(290, 215, duration=0.2)
    pyautogui.write(nfe)
    pyautogui.click(397, 217, duration=0.2)
    pyautogui.write(serie)
    pyautogui.click(447, 218, duration=0.2)
    pyautogui.write(modelo)
    pyautogui.press('enter')
    pyautogui.click(438, 88, duration=0.2)
    pyautogui.click(749, 480, duration=0.2)
    pyautogui.write(cnpj)
    pyautogui.click(862, 431, duration=0.2)
    pyautogui.doubleClick(334, 258)
    pyautogui.write(data)
    pyautogui.doubleClick(430, 256)
    pyautogui.write(data)
    pyautogui.click(569, 257, duration=0.2)
    pyautogui.write(operacao)
    pyautogui.press('enter')
    pyautogui.click(1561, 90, duration=0.2)
    sleep(3)
    pyautogui.click(1485, 289, duration=0.2)
    sleep(2)

def imprimir_item_nota(row):
    pyautogui.press('enter')
    descricao_item = str(row['Descrição'])
    qnt = str(row['Qnt'])
    total = str(row['Total'])
    placa = "*" + str(row['Placa'])

    if "ARLA" in descricao_item:
        descricao_item = "*arla"
    elif "S10" in descricao_item:
        descricao_item = "*s10"
    elif "S500" in descricao_item:
        descricao_item = "*s500"
    elif "gasolina" in descricao_item:
        descricao_item = "*gasolina"

    pyautogui.write(descricao_item)
    pyautogui.press('enter')

    if "*arla" in descricao_item:
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.press('enter')

    if "*gasolina" in descricao_item:
        pyautogui.press('down')
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.press('enter')

    if "*s10" in descricao_item or "*s500" in descricao_item:
        pyautogui.press('enter')
    pyautogui.write(qnt)
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.write(total)
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.write(placa)

    if placa == "*CCU7C57" or "*BRY4I72":
        pyautogui.press('enter')

    pyautogui.press('enter')
    pyautogui.click(1556, 284, duration=0.2)
    sleep(3)

#FUNÇÕES GLOBAIS
def limpar_pasta_notas():
    print("limpar_pasta_notas(Global)")
    diretorio_notas = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas"
    for arquivo in os.listdir(diretorio_notas):
        caminho_arquivo = os.path.join(diretorio_notas, arquivo)
        try:
            if os.path.isfile(caminho_arquivo):
                os.remove(caminho_arquivo)
            elif os.path.isdir(caminho_arquivo):
                shutil.rmtree(caminho_arquivo)
        except PermissionError as e:
            print(f"Erro de permissão ao excluir {caminho_arquivo}: {e}")

def formatar_cnpj(cnpj):
    print("formatar_cnpj(Global)")
    cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    return cnpj_formatado

def processar_xml(caminho_arquivo_xml, planilha_de_placas):
    print("processar_xml(caminho_arquivo_xml, planilha_de_placas)(Global)")
    tree = ET.parse(caminho_arquivo_xml)
    root = tree.getroot()
    ns = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

    resultados = {}

    for prod in root.findall(".//nfe:det", namespaces=ns):
        descricao = prod.find(".//nfe:xProd", namespaces=ns).text
        qnt = float(prod.find(".//nfe:qCom", namespaces=ns).text.replace(',', '.'))
        r_l = float(prod.find(".//nfe:vUnCom", namespaces=ns).text.replace(',', '.'))
        total = float(prod.find(".//nfe:vProd", namespaces=ns).text.replace(',', '.'))

        if "ARLA" in descricao:
            descricao = "ARLA"
        
        if "gasolina" in descricao or "GASOLINA" in descricao:
            descricao = "gasolina"

        if "S10" in descricao or "S-10" in descricao or "S DEZ" in descricao or "S 10" in descricao or "S 1 0" in descricao:
            descricao = "S10"

        if "S500" in descricao or "OLEO DIESEL COMUM" in descricao or "S 500" in descricao or "S-500" in descricao:
            descricao = "S500"

        observacoes = root.find(".//nfe:infAdic/nfe:infCpl", namespaces=ns).text
        placa_match = re.search(r'[A-Z]{3}-?\d[A-Z0-9]-?\d{2}', observacoes)

        if placa_match:
            placa_encontrada = placa_match.group().upper().replace("-", "")
            if placa_encontrada in planilha_de_placas:
                placa_resultado = placa_encontrada
            else:
                placa_resultado = ("Placa não encontrada na planilha")
        else:
            placa_resultado = ("Placa não encontrada no campo de observações")

        if descricao in resultados:
            resultados[descricao]['qnt'] += qnt
            resultados[descricao]['r_l'] += r_l
            resultados[descricao]['total'] += total
        else:
            resultados[descricao] = {'qnt': qnt, 'r_l': r_l, 'total': total, 'placa': placa_resultado}

    lista_resultados = [
        (
            root.find(".//nfe:ide/nfe:nNF", namespaces=ns).text,
            root.find(".//nfe:ide/nfe:serie", namespaces=ns).text,
            formatar_cnpj(root.find(".//nfe:emit/nfe:CNPJ", namespaces=ns).text),
            parser.isoparse(root.find(".//nfe:ide/nfe:dhEmi", namespaces=ns).text).strftime("%d/%m/%Y"),
            descricao,
            f'{resultados[descricao]["qnt"]:.2f}'.replace('.', ','),
            f'{resultados[descricao]["r_l"]:.2f}'.replace('.', ','),
            f'{resultados[descricao]["total"]:.2f}'.replace('.', ','),
            resultados[descricao]['placa'],
            root.find(".//nfe:emit/nfe:xNome", namespaces=ns).text,
            formatar_cnpj(root.find(".//nfe:infNFe/nfe:dest/nfe:CNPJ", namespaces=ns).text),
        ) for descricao in resultados
    ]
  
    # Adicionar cabeçalhos às listas de resultados
    lista_resultados_com_cabecalhos = lista_resultados
    return lista_resultados_com_cabecalhos
caminho_arquivo_notas = r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\notas.xlsx"
planilha_de_placas = pd.read_excel(r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\placas.xlsx")['Placa'].tolist()

#FUNÇÕES RATEIO
def buscar_fechamento_rateio(codigo_find):
    print("Buscando Fechamento Rateio")
    limpar_pasta_notas() #FUNÇÃO GLOBAL
    # Obter o HTML da página
    html_content = driver.page_source
    # Parsear o HTML
    tree = html.fromstring(html_content)
    fechamento_encontrado = False
    for i in range(1, 11):
        codigo_elemento = tree.xpath('//*[@id="root"]/main/div[2]/div[2]/div[1]/div/div/table[{}]/thead/tr/th[1]'.format(i))
        if codigo_elemento:
            codigo_valor = codigo_elemento[0].text.strip() if codigo_elemento[0].text else None
            # Aqui é onde a modificação ocorre, apenas a parte numérica é mantida
            codigo_Número = codigo_valor.split(": ")[1].split(")")[0] if codigo_valor else None
            if str(codigo_find) == codigo_Número:
                xml_botao = '//*[@id="root"]/main/div[2]/div[2]/div[1]/div/div/table[{}]/tbody/tr/td[1]/div/div[7]/button'.format(i)
                # Clicar no botão para baixar xml
                botao_xml = driver.find_element(By.XPATH, xml_botao)
                botao_xml.click()
                sleep(3)
                fechamento_encontrado = True
                pasta_xmls = r"C:\Users\Log20-2\Desktop\Versão 2.0\Notas"
                resultados_totais = []  # Lista para armazenar todos os resultados processados

                for nome_arquivo in os.listdir(pasta_xmls):
                    caminho_xml = os.path.join(pasta_xmls, nome_arquivo)
                    resultados = processar_xml(caminho_xml, planilha_de_placas)
                    resultados_totais.extend(resultados)  # Adiciona os resultados ao total

                # Criar um DataFrame com todos os resultados
                df_resultados = pd.DataFrame(resultados_totais)

                # Adicionar os cabeçalhos ao DataFrame
                df_resultados.columns = ["Número", "Série", "Emitente CNPJ", "Emissão", "Descrição", "Qnt", "Reais/L", "Total", "Placa", "Emitente", "Filial"]

                # Salvar o DataFrame em um arquivo XLSX
                caminho_arquivo_xlsx = r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\notas.xlsx"

                df_resultados.to_excel(caminho_arquivo_xlsx, index=False)
                
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(["NF", "Prod", "ICMS"])

                for icms in resultados_totais:
                    if "S10" in icms[4] or "S500" in icms[4]:
                        # Fazer a multiplicação do elemento 5 por 1.0635
                        nf_icms = float(icms[0])
                        prod_icms = (icms[4])
                        qnt_diesel = float(icms[5].replace(',', '.'))
                        resultado_icms = qnt_diesel * 1.0635
                        resultado_icms = round(resultado_icms, 2)  # Arredonda para 2 casas decimais
                        sheet.append([nf_icms, prod_icms, resultado_icms])
                        print(resultado_icms)
                caminho_icms = os.path.join(os.getcwd(), r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\icms.xlsx")
                workbook.save(caminho_icms)

                filial = resultados_totais[-1]
                mensagem = f"{filial[10]}"
                messagebox.showwarning("Lançar na Filial:",mensagem)

                # Clicar no botão para baixar os abastecimentos
                botao_abastecimentos = '//*[@id="root"]/main/div[2]/div[2]/div[1]/div/div/table[{}]/tbody/tr/td[1]/div/div[15]/button[1]'.format(i)
                botao_abastecimentos = driver.find_element(By.XPATH,botao_abastecimentos)
                botao_abastecimentos.click()
                sleep(3)
                salvar_porcentagens(arquivo_csv,caminho_salvamento)
                
    if not fechamento_encontrado:
        messagebox.showinfo("Procurando Fechamento no Gasola","Fechamento não entrado")

def salvar_porcentagens(arquivo_csv, arquivo_excel):
    # Leitura do arquivo CSV usando ponto e vírgula como separador
    df = pd.read_csv(arquivo_csv, encoding='utf-8', sep=';')

    # Renomeia as colunas para remover espaços e caracteres especiais
    df.columns = df.columns.str.strip().str.replace(' ', '_').str.replace('/', '_')

    # Imprime as colunas do DataFrame
    print("Colunas do DataFrame:")

    try:
        # Substituir ',' por '.' e converter para numérico
        df['Valor_Pago'] = pd.to_numeric(df['Valor_Pago'].str.replace(',', '.'), errors='coerce')
        df['Litros'] = pd.to_numeric(df['Litros'].str.replace(',', '.'), errors='coerce')

        # Criação da tabela dinâmica
        tabela_dinamica = pd.pivot_table(df, values='Valor_Pago', index='Placa', columns='Combustível', aggfunc='sum', fill_value=0, margins=True, margins_name='Total Geral')

        # Calcular porcentagens em relação ao total geral
        tabela_dinamica_porcentagens = tabela_dinamica.div(tabela_dinamica.iloc[-1, :], axis=1) * 100
        tabela_dinamica_porcentagens = tabela_dinamica_porcentagens.iloc[:-1, :]  # Excluir a última linha 'Total Geral'

        # Ajustar o formato das células na tabela de porcentagens
        tabela_dinamica_porcentagens = tabela_dinamica_porcentagens.round(2)

        # Corrigir porcentagens para garantir que a soma seja exatamente 100%
        tabela_dinamica_porcentagens = tabela_dinamica_porcentagens * (100 / tabela_dinamica_porcentagens.sum())

        # Escrever a tabela dinâmica e porcentagens em um arquivo Excel
        with pd.ExcelWriter(arquivo_excel, engine='xlsxwriter') as writer:
            tabela_dinamica_porcentagens.to_excel(writer, sheet_name='Porcentagens', index=True)

            # Obter as folhas do Excel
            workbook = writer.book
            worksheet_p = writer.sheets['Porcentagens']

            # Formato padrão para as células
            formato_padrao = workbook.add_format({'bold': False, 'num_format': '#,##0.00'})

            # Formato para porcentagens com duas casas decimais
            formato_porcentagem = workbook.add_format({'num_format': '0.00%'})

            # Aplicar formato padrão a todas as células em ambas as folhas
            for worksheet in [worksheet_p]:
                for col_num, value in enumerate(tabela_dinamica.columns.get_level_values('Combustível')):
                    worksheet.set_column(col_num, col_num, None, formato_padrao)

                if 'Porcentagens' in worksheet.name:
                    for col_num, value in enumerate(tabela_dinamica_porcentagens.columns):
                        worksheet.set_column(col_num + len(tabela_dinamica.columns), col_num + len(tabela_dinamica.columns), None, formato_porcentagem)

        # Imprime a tabela dinâmica e porcentagens
        print(tabela_dinamica)
        
        print("\nTabela com Porcentagens:")
        print(tabela_dinamica_porcentagens)
        print(f"\nTabelas criadas e salvas em {arquivo_excel}")
        # Contar o número de placas na planilha
        numero_de_placas = len(tabela_dinamica_porcentagens.index)

        # Imprimir o número de placas
        if numero_de_placas <=1:
            messagebox.showwarning("Preparando rateio","ATENÇÃO!!!\nEste fechamento é notaplaca")
        else:
            messagebox.showinfo("Prearando rateio","Fechamento pronto para ser lançado")
    except Exception as e:
        messagebox.showerror("Erro ao criar a tabela dinâmica","Ops, este fechamento está com problemas\nInfelizmente terá que fazer o lançamento manualmente")

#FUNÇÕES LANÇAMENTO RATEIO
def rateio():
    janela_rateio()
    print_header = True
    # Lendo o arquivo Excel
    caminho_excel = r"C:\Users\Log20-2\Desktop\Versão 2.0\planilhas\notas.xlsx"
    df = pd.read_excel(caminho_excel)

    # Agrupar por informação da nota fiscal
    df_agrupado = df.groupby(['Número', 'Série', 'Emitente CNPJ', 'Emissão', 'Descrição'])

    for chave, row in df.iterrows():
        nfe = row['Número']
        serie = row['Série']
        modelo = "55"
        cnpj = row['Emitente CNPJ']
        data = str(row['Emissão']).replace('/','')
        operacao = "5015"

        pyautogui.click(1492, 89, duration=0.2)
        pyautogui.click(290, 215, duration=0.2)
        pyautogui.write(str(nfe))
        pyautogui.click(397, 217, duration=0.2)
        pyautogui.write(str(serie))
        pyautogui.click(447, 218, duration=0.2)
        pyautogui.write(modelo)
        pyautogui.press('enter')
        pyautogui.click(438, 88, duration=0.2)
        pyautogui.click(749, 480, duration=0.2)
        pyautogui.write(str(cnpj))
        pyautogui.click(862, 431, duration=0.2)
        pyautogui.doubleClick(334, 258)
        pyautogui.write(data)
        pyautogui.doubleClick(430, 256)
        pyautogui.write(data)
        pyautogui.click(569, 257, duration=0.2)
        pyautogui.write(operacao)
        pyautogui.press('enter')
        pyautogui.click(1561, 90, duration=0.2)
        sleep(3)
        pyautogui.click(1485, 289, duration=0.2)
        sleep(2)

        if print_header:
            print_header = False  
        for index, row in df.iterrows():
            pyautogui.press('enter')
            item = "*" + str(row['Descrição'])
            print(item)
            qnt = str(row['Qnt'])
            print(qnt)
            total = str(row['Total'])
            print(total)
            placa = "*rky6h45"
            print(f"Item: {item}")

            if "ARLA" in item:
                item = "*arla"
                pyautogui.write(item)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(qnt)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(total)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(placa)
                pyautogui.press('enter')         
                pyautogui.click(1556,284,duration=0.2)  
                sleep(2)              
        
            elif "S10" in item:
                item = "*s10"
                pyautogui.write(item)
                pyautogui.press('enter')   
                pyautogui.press('enter')
                pyautogui.write(qnt)
                pyautogui.press('enter')
                pyautogui.press('enter')       
                pyautogui.write(total)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(placa)
                pyautogui.press('enter')         
                pyautogui.click(1556,284,duration=0.2)    
                sleep(2)

            elif "S500" in item or "DIESEL COMUM" in item:
                item = "*s500"
                pyautogui.write(item)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(qnt)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(total)
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.write(placa)
                pyautogui.press('enter')
                pyautogui.click(1556,284,duration=0.2)
                sleep(2)
        break
    pyautogui.press('esc')

def arla():
    pyautogui.click(1068,508)
    pyautogui.click(769,499)
    excel_path = r'C:\Users\Log20-2\Desktop\Versão 2.0\Rateio\abastecimentos.xlsx'
    df = pd.read_excel(excel_path)

    # Flag para imprimir o cabeçalho apenas uma vez
    print_header = True

    for index, row  in df.iterrows():
        placa_arla = "*" + (row['Placa'])  
        porcentagem = "{:.3f}".format(row['Arla Granel'])
        print(placa_arla)
        print (porcentagem)
        if porcentagem == '0.000':
            continue
        if porcentagem == '100.0':
            if porcentagem == '0.000':
                continue
            pyautogui.click(1044, 511)
            pyautogui.click(893, 580)
            pyautogui.write(placa_arla)
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.doubleClick(829, 662)
            pyautogui.click(1117, 510)
            break
        pyautogui.click(1044, 511)
        pyautogui.click(893,580)
        pyautogui.write(placa_arla)
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.doubleClick(829, 662)
        pyautogui.write(porcentagem)
        pyautogui.click(1117, 510)

def s500():
    pyautogui.click(1068,508)
    pyautogui.click(769,499)
    excel_path = r'C:\Users\Log20-2\Desktop\Versão 2.0\Rateio\abastecimentos.xlsx'
    df = pd.read_excel(excel_path)

    # Flag para imprimir o cabeçalho apenas uma vez
    print_header = True

    for index, row  in df.iterrows():
        placa_s500 = "*" + (row['Placa'])  
        porcentagem ="{:.3f}".format(row['Diesel S500'])
        print(placa_s500)
        print (porcentagem)
        if porcentagem == '0.000':
            continue
        if porcentagem == '100.0':
            if porcentagem == '0.000':
                continue
            pyautogui.click(1044, 511)
            pyautogui.click(893, 580)
            pyautogui.write(placa_s500)
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.doubleClick(829, 662)
            pyautogui.click(1117, 510)
            break
        pyautogui.click(1044, 511)
        pyautogui.click(893,580)
        pyautogui.write(placa_s500)
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.doubleClick(829, 662)
        pyautogui.write(porcentagem)
        pyautogui.click(1117, 510)
        if keyboard.is_pressed('q'):
            break

def s10():
    pyautogui.click(1068,508)
    pyautogui.click(769,499)
    excel_path = r'C:\Users\Log20-2\Desktop\Versão 2.0\Rateio\abastecimentos.xlsx'
    df = pd.read_excel(excel_path)

    # Flag para imprimir o cabeçalho apenas uma vez
    print_header = True

    for index, row  in df.iterrows():
        placa_s10 = "*" + (row['Placa'])  
        porcentagem = "{:.3f}".format(row['Diesel S10'])
        print(placa_s10)
        print(porcentagem)
        if porcentagem == '0.000':
            continue
        if porcentagem == '100.0':
            if porcentagem == '0.000':
                continue
            pyautogui.click(1044, 511)
            pyautogui.click(893, 580)
            pyautogui.write(placa_s10)
            pyautogui.press('enter')
            pyautogui.press('enter')
            pyautogui.doubleClick(829, 662)
            pyautogui.click(1117, 510)
            break
        pyautogui.click(1044, 511)
        pyautogui.click(893,580)
        pyautogui.write(placa_s10)
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.doubleClick(829, 662)
        pyautogui.write(porcentagem)
        pyautogui.click(1117, 510)
            
        if keyboard.is_pressed('q'):
            break

def janela_rateio():
    janela_do_rateio = Tk()
    janela_do_rateio.title("Sistema de Lançamentos")
    janela_do_rateio.geometry("600x180")
    janela_do_rateio.configure(background="#ffffff")
    botao_arla=Button(janela_do_rateio, text="Arla",command=arla,background="#ffffff")
    botao_arla.config(font=("Arial", 16))
    botao_arla.place(x=20, y=25, width=170, height=70)
    botao_s10=Button(janela_do_rateio, text="S10",command=s10,background="#ffffff")
    botao_s10.config(font=("Arial", 16))
    botao_s10.place(x=234, y=25, width=170, height=70)
    botao_500=Button(janela_do_rateio, text="S500",command=s500,background="#ffffff")
    botao_500.config(font=("Arial", 16))
    botao_500.place(x=444, y=25, width=170, height=70)

def formatar_numero_nf(numero_nf):
    numero_nf_str = str(numero_nf).zfill(9)
    numero_nf_formatado = '.'.join(numero_nf_str[i:i+3] for i in range(0, 9, 3))
    return numero_nf_formatado

  
    

diretorio_projeto = r"C:\\Users\\Log20-2\\Desktop\\Versão 2.0"


janela = Tk()
janela.title("Sistema de Lançamentos")
janela.geometry("800x500")
janela.configure(background="#ffffff")

tree = ttk.Treeview(janela, columns=("valores_codigo", "valores_nome", "valorp_posto","valor_periodo_inicio","valor_periodo_fim","valor_quantidade_arquivos",))
tree.heading("valores_codigo", text="Fechamento")
tree.heading("valores_nome", text="Posto")
tree.heading("valorp_posto", text="Valor")
tree.heading("valor_periodo_inicio", text="Inicio")
tree.heading("valor_periodo_fim", text="Fim")
tree.heading("valor_quantidade_arquivos", text = "Arquivos")
tree.column("valores_codigo", width=80)
tree.column("valorp_posto", width=80)
tree.column("valor_periodo_inicio", width=80)
tree.column("valor_periodo_fim", width=80)
tree.place(x=1, y=1, width=800, height=500)
tree["show"] = "headings"
tree.bind("<Button-3>", popup_menu)

acessar_site()
atualizar_tabela()
centralizar_janela(janela)
janela.mainloop()

while True:
    pass

# Copyright (c) 2024 Bernardo Augusto Borges Vilbert
# Este código é propriedade de Bernardo Augusto Borges Vilbert.